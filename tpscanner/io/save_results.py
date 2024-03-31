"""Module to save the results to an Excel file."""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Color, Font, NamedStyle

from tpscanner.config import config
from tpscanner.logger import logger


def save_individual_deals(filename, individual_deals_items) -> None:
    """Save the individual deals to an Excel file.

    Arguments:
        filename (str): The name of the Excel file.
        individual_deals_items (dict): The dictionary of individual deals.

    """
    headers = [
        "Seller",
        "Reviews",
        "Rating",
        "Price",
        "Quantity",
        "Delivery Price",
        "Free Delivery From",
        "Total Price",
        "Total Price + Delivery",
        "Availability",
        "See Offer",
    ]
    keys = [
        "seller",
        "seller_reviews",
        "seller_rating",
        "price",
        "quantity",
        "delivery_price",
        "free_delivery",
        "total_price",
        "total_price_plus_delivery",
        "availability",
        "link",
    ]
    col_format_start_range = 4
    for name, items in individual_deals_items.items():
        logger.info(f"Saving deals for `{name}`.")
        _create_workbook(filename, name, headers, items, keys, col_format_start_range)


def save_best_individual_deals(filename, sheetname, best_deals_items) -> None:
    """Save the best individual deals to an Excel file.

    Arguments:
        filename (str): The name of the Excel file.
        sheetname (str): The name of the sheet.
        best_deals_items (list): The list of best individual deals.

    """
    headers = [
        "Product",
        "Seller",
        "Reviews",
        "Rating",
        "Price",
        "Quantity",
        "Delivery Price",
        "Free Delivery From",
        "Total Price",
        "Total Price + Delivery",
        "Availability",
        "See Offer",
    ]
    keys = [
        "name",
        "seller",
        "seller_reviews",
        "seller_rating",
        "price",
        "quantity",
        "delivery_price",
        "free_delivery",
        "total_price",
        "total_price_plus_delivery",
        "availability",
        "link",
    ]
    col_format_start_range = 4
    _create_workbook(
        filename, sheetname, headers, best_deals_items, keys, col_format_start_range
    )


def save_best_cumulative_deals(filename, sheetname, best_deals_items) -> None:
    """Save the best cumulative deals to an Excel file.

    Arguments:
        filename (str): The name of the Excel file.
        sheetname (str): The name of the sheet.
        best_deals_items (list): The list of best cumulative deals.

    """
    headers = [
        "Seller",
        "Reviews",
        "Rating",
        "Cumulative Price",
        "Delivery Price",
        "Free Delivery from",
        "Cumulative Price + Delivery",
        "Availability",
    ]
    keys = [
        "seller",
        "seller_reviews",
        "seller_rating",
        "cumulative_price",
        "delivery_price",
        "free_delivery",
        "cumulative_price_plus_delivery",
        "availability",
    ]
    col_format_start_range = 3
    _create_workbook(
        filename, sheetname, headers, best_deals_items, keys, col_format_start_range
    )


def _create_workbook(filename, sheetname, headers, items, keys, col_format_start_range):
    filename = os.path.join(config.output_dir, filename)
    # ensure that the sheet name is less than 31 characters
    if len(sheetname) > 31:
        sheetname = sheetname[:31]
    # Check if the file already exists
    if os.path.exists(filename):
        logger.info(f"File `{filename}` already exists. Opening it...")
        # If the file exists, open it
        workbook = load_workbook(filename)
        # create a new worksheet
        worksheet = workbook.create_sheet(sheetname)
        # find existing named style
        header_style = workbook._named_styles["header_style"]
    else:
        logger.info(f"File `{filename}` does not exist. Creating it...")
        # else create a workbook
        workbook = Workbook()
        # Use the active sheet as the new sheet
        worksheet = workbook.active
        worksheet.title = sheetname
        # create a new header style
        header_style = NamedStyle(
            name="header_style",
            font=Font(bold=True),
            alignment=Alignment(horizontal="center"),
        )

    # Set column names in the header row
    for i, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=i, value=header)

    # apply the header style with bold and center alignment
    header_row = worksheet[1]
    for cell in header_row:
        cell.style = header_style

    # Write the data
    blue_font = Font(color=Color(rgb="2a65d1"), underline="single")
    for row, item in enumerate(items, start=2):
        for i, key in enumerate(keys, start=1):
            if key == "link":
                cell = worksheet.cell(
                    row=row, column=i, value=f'=HYPERLINK("{item[key]}", "Link")'
                )
                cell.font = blue_font
                cell.alignment = Alignment(horizontal="center")
            elif key == "seller":
                cell = worksheet.cell(
                    row=row,
                    column=i,
                    value=f'=HYPERLINK("{item['seller_link']}", "{item[key]}")',
                )
                cell.font = blue_font
            elif key == "seller_reviews":
                cell = worksheet.cell(
                    row=row,
                    column=i,
                    value=f'=HYPERLINK("{item['seller_reviews_link']}", "{item[key]}")',
                )
                cell.font = blue_font
                cell.alignment = Alignment(horizontal="center")
            else:
                worksheet.cell(row=row, column=i, value=item[key])

    # Set number format for the numeric columns
    for col_num in range(col_format_start_range, len(headers) + 1):
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col_num).number_format = "#,##0.00"

    workbook.save(filename)
