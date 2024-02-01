# save_results.py

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle

from tpscanner.logger import logger


def save_intermediate_results(filename, sheetname, items):
    headers = [
        "Seller",
        "Reviews",
        "Price",
        "Quantity",
        "Delivery Price",
        "Free Delivery From",
        "Total Price",
        "Total Price + Delivery",
        "Availability",
    ]
    keys = [
        "seller",
        "seller_reviews",
        "price",
        "quantity",
        "delivery_price",
        "free_delivery",
        "total_price",
        "total_price_plus_delivery",
        "availability",
    ]
    _create_workbook(filename, sheetname, headers, items, keys)


def save_best_individual_deals(filename, sheetname, best_deals_items):
    headers = [
        "Seller",
        "Reviews",
        "Price",
        "Quantity",
        "Delivery Price",
        "Free Delivery From",
        "Total Price",
        "Total Price + Delivery",
        "Availability",
    ]
    keys = [
        "seller",
        "seller_reviews",
        "price",
        "quantity",
        "delivery_price",
        "free_delivery",
        "total_price",
        "total_price_plus_delivery",
        "availability",
    ]
    _create_workbook(filename, sheetname, headers, best_deals_items, keys)


def save_best_cumulative_deals(filename, sheetname, best_deals_items):
    headers = [
        "Seller",
        "Reviews",
        "Cumulative Price",
        "Delivery Price",
        "Free Delivery from",
        "Cumulative Price + Delivery",
        "Availability",
    ]
    keys = [
        "seller",
        "seller_reviews",
        "cumulative_price",
        "delivery_price",
        "free_delivery",
        "cumulative_price_plus_delivery",
        "availability",
    ]
    _create_workbook(filename, sheetname, headers, best_deals_items, keys)


def _create_workbook(filename, sheetname, headers, items, keys):
    # ensure that the sheet name is less than 31 characters
    if len(sheetname) > 31:
        sheetname = sheetname[:31]
    # Check if the file already exists
    if os.path.exists(filename):
        logger.debug(f"File `{filename}` already exists. Opening it...")
        # If the file exists, open it
        workbook = load_workbook(filename)
        # create a new worksheet
        worksheet = workbook.create_sheet(sheetname)
        # find existing named style
        header_style = workbook._named_styles["header_style"]
    else:
        logger.debug(f"File `{filename}` does not exist. Creating it...")
        # else create a workbook
        workbook = Workbook()
        # Use the active sheet as the new sheet
        worksheet = workbook.active
        worksheet.title = sheetname
        # create a new header style
        header_style = NamedStyle(
            name="header_style",
            font=Font(bold=True),
            alignment=Alignment(horizontal="center"),
        )

    # Set column names in the header row
    for i, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=i, value=header)

    # apply the header style with bold and center alignment
    header_row = worksheet[1]
    for cell in header_row:
        cell.style = header_style

    # Write the data
    for row, item in enumerate(items, start=2):
        for i, key in enumerate(keys, start=1):
            worksheet.cell(row=row, column=i, value=item[key])

    # Set number format for the numeric columns
    for col_num in range(3, len(headers) + 1):
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col_num).number_format = "#,##0.00"

    workbook.save(filename)
