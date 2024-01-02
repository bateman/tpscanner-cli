import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle

def save_intermediate_results(filename, sheetname, items):
    title = sheetname
    # ensure that the sheet name is less than 31 characters
    if len(sheetname) > 31:
        title = sheetname[:31]
    # Check if the file already exists
    if os.path.exists(filename):
        print(f"File {filename} already exists. Opening it...")
        # If the file exists, open it 
        workbook = load_workbook(filename)
        # create a new worksheet 
        worksheet = workbook.create_sheet(title)
        # find existing named style
        header_style = workbook._named_styles['header_style']
    else:
        print(f"File {filename} does not exist. Creating it...")
        # else create a workbook
        workbook = Workbook()
        # Use the active sheet as the new sheet
        worksheet = workbook.active
        worksheet.title = title
        # create a new header style
        header_style = NamedStyle(name='header_style', font=Font(bold=True), alignment=Alignment(horizontal='center'))

    # Set column names in the header row
    worksheet.cell(row=1, column=1, value='Seller')
    worksheet.cell(row=1, column=2, value='Reviews')
    worksheet.cell(row=1, column=3, value='Price')
    worksheet.cell(row=1, column=4, value='Delivery Price')
    worksheet.cell(row=1, column=5, value='Free Delivery From')
    worksheet.cell(row=1, column=6, value='Total Price')
    worksheet.cell(row=1, column=7, value='Availability')
    
    # apply the header style with bold and center alignment
    header_row = worksheet[1]
    for cell in header_row:
        cell.style = header_style

    # Write the data
    for row, item in enumerate(items, start=2):
        worksheet.cell(row=row, column=1, value=item['seller'])
        worksheet.cell(row=row, column=2, value=item['seller_reviews'])
        worksheet.cell(row=row, column=3, value=item['price'])
        worksheet.cell(row=row, column=4, value=item['delivery_price'])
        worksheet.cell(row=row, column=5, value=item['free_delivery'])
        worksheet.cell(row=row, column=6, value=item['total_price'])
        worksheet.cell(row=row, column=7, value=item['availability'])
    
    # Set number format for the 'Price', 'Delivery Price',  'Free Delivery From', 'Total Price' columns
    for col_num in range(3, 7):
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col_num).number_format = '#,##0.00'
            
    workbook.save(filename)
    
    
def save_best_deals(filename, title, best_deals_items):
    workbook = load_workbook(filename)
    # create a new worksheet 
    worksheet = workbook.create_sheet(title)
    # find existing named style
    header_style = workbook._named_styles['header_style']
    
    # Set column names in the header row
    worksheet.cell(row=1, column=1, value='Seller')
    worksheet.cell(row=1, column=2, value='Reviews')
    worksheet.cell(row=1, column=3, value='Cumulative Price')
    worksheet.cell(row=1, column=4, value='Delivery Price')
    worksheet.cell(row=1, column=5, value='Free Delivery from')
    worksheet.cell(row=1, column=6, value='Availability')
    
    # apply the header style with bold and center alignment
    header_row = worksheet[1]
    for cell in header_row:
        cell.style = header_style
        
    # Write the data
    for row, item in enumerate(best_deals_items, start=2):
        worksheet.cell(row=row, column=1, value=item['seller'])
        worksheet.cell(row=row, column=2, value=item['seller_reviews'])
        worksheet.cell(row=row, column=3, value=item['cumulative_price'])
        worksheet.cell(row=row, column=4, value=item['delivery_price'])
        worksheet.cell(row=row, column=5, value=item['free_delivery'])
        worksheet.cell(row=row, column=6, value=item['availability'])
        
    # Set number format for the numeric columns
    for col_num in range(3, 6):
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col_num).number_format = '#,##0.00'
            
    workbook.save(filename)
